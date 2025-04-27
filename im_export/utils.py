import openpyxl
from decimal import Decimal


def parse_bank_file(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    first_rows = [row for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True)]

    if any("Txn. Date" in str(cell) for row in first_rows for cell in row if cell):
        return parse_excel_exim(file)
    else:
        return parse_excel_bdc(file)


# Fonction pour parser les fichier BDC
def parse_excel_bdc(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    transactions = []
    total_kmf = Decimal("0.00")

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if not row or len(row) < 5:
            continue

        try:
            date = row[2]
            amount_raw = row[4]

            # Convertir le montant (gestion du format "2.000,50")
            amount_str = str(amount_raw).replace(",", ".").replace(" ", "")
            amount = Decimal(amount_str)

            if amount > 0:
                transactions.append({
                    "date": str(date),
                    "description": row[5],
                    "amount": str(amount),
                })
                total_kmf += amount
        except Exception as e:
            print(f"Erreur à la ligne {row_idx}: {row} - {e}")
            continue

    return {
        "transactions": transactions,
        "total_kmf": str(total_kmf),
        "count": len(transactions),
    }


# Fonction pour parser les fichier Exim
def parse_excel_exim(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    transactions = []
    total_kmf = Decimal("0.00")
    start_parsing = False

    for row in sheet.iter_rows(values_only=True):
        if not start_parsing:
            if row and any("Txn. Date" in str(cell) for cell in row if cell):
                headers = [str(h).strip() if h else None for h in row]
                try:
                    date_idx = headers.index("Txn. Date")
                    desc_idx = headers.index("Description")
                    credit_idx = headers.index("Credit")
                except ValueError as e:
                    raise Exception("Colonnes attendues non trouvées dans le fichier (Txn. Date, Description, Credit)") from e
                start_parsing = True
                continue

        if start_parsing:
            if row and str(row[1]).startswith("Opening Balance"):
                break

            try:
                credit_val = row[credit_idx]
                if credit_val and Decimal(credit_val) > 0:
                    transactions.append({
                        "date": str(row[date_idx]),
                        "description": row[desc_idx],
                        "amount": str(Decimal(credit_val)),
                    })
                    total_kmf += Decimal(credit_val)
            except Exception as e:
                print(f"Erreur parsing ligne: {row} - {e}")
                continue

    return {
        "transactions": transactions,
        "total_kmf": str(total_kmf),
        "count": len(transactions),
    }
