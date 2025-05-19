import logging
from typing import Tuple, Any, Dict, List
from tablib import Dataset

from im_export.resources import InsureeResource
import openpyxl
from decimal import Decimal
from invoice.models import Invoice, PaymentInvoice, DetailPaymentInvoice, InvoiceEvent
from insuree.models import Insuree, Family
from insuree.services import FamilyService, InsureeService
from contribution.models import Premium
from policy.models import Policy
from payer.models import Payer
from datetime import datetime
from uuid import uuid4
from django.contrib.contenttypes.models import ContentType
from contribution.services import update_or_create_premium
from collections import defaultdict
from location.models import Location
from django.db.models import Q
from datetime import datetime as py_datetime
from core.datetimes.shared import datetimedelta
from contribution_plan.models import ContributionPlan
from django.db import transaction

logger = logging.getLogger(__name__)


class InsureeImportExportService:
    supported_content_types = {
        'xls': 'application/vnd.ms-excel',
        'csv': 'text/csv',
        'json': 'application/json',
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    }

    class Strategy:
        INSERT = 'INSERT'
        UPDATE = 'UPDATE'
        INSERT_UPDATE = "INSERT_UPDATE"

    supported_strategies = (Strategy.INSERT, Strategy.UPDATE, Strategy.INSERT_UPDATE)

    def __init__(self, user):
        self._user = user
        self._resource = InsureeResource(user)

    def export_insurees(self, export_format: str = 'csv') -> Tuple[str, Any]:
        if export_format not in self.supported_content_types:
            raise ValueError(f'Non-supported export format: {export_format}')

        # All supported formats match Tablib attrs, to update if that's not valid anymore
        return self.supported_content_types[export_format], \
            getattr(self._resource.export(), export_format)

    def import_insurees(self, import_file, dry_run: bool = False, strategy: str = Strategy.INSERT) \
            -> Tuple[bool, Dict[str, int], List[str]]:

        if not import_file:
            return self._get_general_error('Missing import file')
        if strategy not in self.supported_strategies:
            return self._get_general_error(f'Non-supported strategy: {strategy}')

        # Other strategies are not supported for now
        if strategy in (InsureeImportExportService.Strategy.UPDATE, InsureeImportExportService.Strategy.INSERT_UPDATE):
            strategy = InsureeImportExportService.Strategy.INSERT
            logger.warning(f'Strategy {strategy} not currently supported, defaulting to {InsureeImportExportService.Strategy.INSERT}')

        try:
            if import_file.content_type == 'application/vnd.ms-excel':
                data_set = Dataset(headers=InsureeResource.insuree_headers).load(import_file.open(), 'xls')
            elif import_file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                data_set = Dataset(headers=InsureeResource.insuree_headers).load(import_file.open(), 'xlsx')
            elif import_file.content_type == 'application/json':
                data_set = Dataset(headers=InsureeResource.insuree_headers).load(import_file.open())
            else:
                data_set = Dataset(headers=InsureeResource.insuree_headers).load(import_file.read().decode())
        except Exception as e:
            return self._get_general_error('Failed to parse input file', e)

        try:
            data_set = self._resource.validate_and_sort_dataset(data_set)
        except Exception as e:
            return self._get_general_error('file validation failed', e)

        dry_run_result = self._resource.import_data(data_set, dry_run=True)  # Test the data import
        totals = self._get_totals_from_result(dry_run_result)
        errors = self._get_errors_from_result(dry_run_result)
        success = not dry_run_result.has_errors() and not dry_run_result.has_validation_errors()

        if not dry_run:
            if success:
                self._resource.import_data(data_set, dry_run=False)  # Actually import
                logger.info(f'Imported {totals["sent"]} insurees')
            else:
                logger.info(f'Failed to import {totals["sent"]} insurees, details: {totals}, errors: {errors}')

        return success, totals, errors

    @staticmethod
    def _get_totals_from_result(result):
        return {
            'sent': result.total_rows,
            'created': result.totals['new'],
            'updated': result.totals['update'],
            'deleted': result.totals['delete'],
            'skipped': result.totals['skip'],
            'invalid': result.totals['invalid'],
            'failed': result.totals['error']
        }
        
    @staticmethod
    def _get_general_error(*args):
        errors = []
        for arg in args:
            errors.append(arg.message if hasattr(arg, 'message') else str(arg))
        totals = {'sent': 0, 'created': 0, 'updated': 0, 'deleted': 0, 'skipped': 0, 'invalid': 0, 'failed': 0}
        success = False

        return success, totals, errors

    @staticmethod
    def _get_errors_from_result(result):
        errors = []
        if result.has_validation_errors():
            for invalid_row in result.invalid_rows:
                errors.append(f"row ({invalid_row.number}) - {invalid_row.error.messages}")
        if result.has_errors():
            for index, row_error in result.row_errors():
                for error in row_error:
                    errors.append(f"row ({index}) - {error.error}")
        return errors


class FamilyImportExportService:
    supported_content_types = {
        'xls': 'application/vnd.ms-excel',
        'csv': 'text/csv',
        'json': 'application/json',
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    }

    class Strategy:
        INSERT = 'INSERT'
        UPDATE = 'UPDATE'
        INSERT_UPDATE = "INSERT_UPDATE"

    supported_strategies = (Strategy.INSERT, Strategy.UPDATE, Strategy.INSERT_UPDATE)

    def __init__(self, user):
        self._user = user
        # self._resource = InsureeResource(user)

    def export_families(self, export_format: str = 'csv') -> Tuple[str, Any]:
        if export_format not in self.supported_content_types:
            raise ValueError(f'Non-supported export format: {export_format}')

        # All supported formats match Tablib attrs, to update if that's not valid anymore
        return self.supported_content_types[export_format], \
            getattr(self._resource.export(), export_format)

    def import_families(self, import_file, dry_run: bool = False, strategy: str = Strategy.INSERT) \
            -> Tuple[bool, Dict[str, int], List[str]]:

        if not import_file:
            return InsureeImportExportService._get_general_error('Missing import file')
        if strategy not in self.supported_strategies:
            return InsureeImportExportService._get_general_error(f'Non-supported strategy: {strategy}')

        # Other strategies are not supported for now
        if strategy in (InsureeImportExportService.Strategy.UPDATE, InsureeImportExportService.Strategy.INSERT_UPDATE):
            strategy = InsureeImportExportService.Strategy.INSERT
            logger.warning(f'Strategy {strategy} not currently supported, defaulting to {InsureeImportExportService.Strategy.INSERT}')

        family_headers = ['Identification', 'Etatmatrimonial', 'Membresménage',
       'Nom&prénom_membresménag', 'Sexe', 'Lien de parenté', "Pièce d'identité",
       'NIN', 'Jour de naissance', 'Mois de naissance', 'Année de naissance', 'Âge',
       'Formation ou non', 'Types de formation', 'Maladie invalidante Non',
       'Handicap Non', 'Couverture_Assurance_Mutuelle',
       'Catégories_professionnelles', 'Tailleménages', 'Scores_taille_des_ménages',
       'Revenus', 'Scores_revenus', 'Scores_types_habitation',
       'Scores_totaux_catégorisation', 'Cotisationsfamilles', 'Cotisations_famille_chef_famille',
       'île', 'milieu de résidence', 'District_sanitaire', 'Commune', 'Localité',
       'Taillefamille', 'Autresménage', 'Féminin', 'Masculin',
       'Cotisationsautresménages', 'Cotisations_totales_ménages',
       'PartsGouv_&_PTFCotisations Familles',
       'PartsGouv_&_PTFCotisations_Autresménages_démunis',
       'Parts Gouv & PTFCotisations_Autresménages_Vulnérables',
       'Partstotaux_Gvt&PTF']
        # Education, Relations, Profession, changer les chiffres de income levels
        try:
            if import_file.content_type == 'application/vnd.ms-excel':
                data_set = Dataset(headers=family_headers).load(import_file.open(), 'xls')
            elif import_file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                data_set = Dataset(headers=family_headers).load(import_file.open(), 'xlsx')
            elif import_file.content_type == 'application/json':
                data_set = Dataset(headers=family_headers).load(import_file.open())
            else:
                data_set = Dataset(headers=family_headers).load(import_file.read().decode())
        except Exception as e:
            return InsureeImportExportService._get_general_error('Failed to parse input file', e)
        gender_dict = {
            1: "M",
            2: "F",
            3: "O"
        }
        professional_situations = {
            0: "Sans profession",
            1: "Agriculteur exploitant/Pêcheur/Éleveur/ou Artisan",
            2: "Commerçants et assimilés",
            3: "Chef d'entreprise",
            4: "Professionnel libéral",
            5: "Enseignant",
            6: "Enseignant chercheur",
            7: "Médecin",
            8: "Autre professionnel de Santé",
            9: "Autorité de l'admin. publique",
            10: "Employé de l'administration publique",
            11: "Employé administratif ou commercial d'entreprise",
            12: "Ingénieur et Technicien d'entreprise",
            13: "Ingénieur et Technicien travaillant à compte propre",
            14: "Clergé, religieux",
            15: "Policier et militaire",
            16: "Personnel de services directs aux particuliers",
            17: "Artiste",
            18: "Ouvrier",
            19: "Retraité"
        }
        grouped = defaultdict(list)
        today = py_datetime.today()
        for row in data_set.dict:
            grouped[row['Identification']].append(row)
        try:
            familycreated = 0
            with transaction.atomic():
                for identification, rows in grouped.items():
                    logger.info(f"Memmbers for Identification = {identification}:")
                    # Mise par Ordre de membre_de_menage ascendant en commencant par le chef de famille
                    sorted_members = sorted(rows, key=lambda x: int(x['Membresménage']) if x['Membresménage'] not in [None, ""] else -1)
                    parent_family = None
                    for r in sorted_members:
                        logger.info(parent_family)
                        yob = r.get("Année de naissance")
                        if yob is None or yob == "-999999999" or len(str(yob)) != 4 or yob == "":
                            if r.get("Lien de parenté") != "" and r.get("Lien de parenté") is not None:
                                if int(r.get("Lien de parenté")) == 3:
                                    years = 10
                                elif int(r.get("Lien de parenté")) in [1, 2, 25]:
                                    years = 30
                                elif int(r.get("Lien de parenté")) in [21, 22, 23, 24]:
                                    years = 40
                                else:
                                    years = 50
                                yob = str(today - datetimedelta(
                                    years=years
                                )).split(" ")[0].split("-")[0]
                            else:
                                yob = "1990"
                        if r.get("Mois de naissance") is None or str(r.get("Mois de naissance")) in ["-999999999", ""]:
                            mob = "01"
                        else:
                            mob = str(r.get("Mois de naissance"))
                            mob = mob.zfill(2)# 1 becomes 01
                            if len(mob) != 2:
                                mob = "01"
                        if r.get("Jour de naissance") is None or str(r.get("Jour de naissance")) in ["-999999999", ""]:
                            dob = "01"
                        else:
                            dob = str(r.get("Jour de naissance"))
                            dob = dob.zfill(2) # 1 becomes 01
                            if len(dob) != 2:
                                dob = "01"
                        village = r.get("Localité")
                        if village == "" or village is None:
                            village = 1
                        current_village_id = Location.objects.filter(
                            Q(code=village) | Q(name=village)).filter(
                                validity_to__isnull=True, type='V').first()
                        if not current_village_id:
                            current_village_id = Location.objects.filter(
                                validity_to__isnull=True, type='V').first().id
                        current_gender = gender_dict.get(3)
                        if r.get("Sexe") is not None and int(r.get("Sexe")) in [1, 2]:
                            current_gender = gender_dict.get(int(r.get("Sexe")))
                        nin_ok = False
                        nin = False
                        if r.get("NIN") is not None and r.get("NIN") != "":
                            nin = str(r.get("NIN")).replace(" ", "")
                            if len(nin) == 7 or (len(nin)==9 and nin.startswith("UG")):
                                nin_ok = True
                        card_issued = True
                        if r.get("Pièce d'identité") is not None and r.get("Pièce d'identité") != "":
                            if not nin_ok or int(r.get("Pièce d'identité")) != 1:
                                card_issued = False
                        marital = r.get("Etatmatrimonial")
                        head = False
                        if marital is not None and marital != "":
                            marital = int(marital)
                            head = True if marital != 2 else False
                        else:
                            marital = 1
                        head_insuree_data = {
                            "last_name": r.get("Nom&prénom_membresménag") if r.get("Nom&prénom_membresménag")\
                                is not None else " ",
                            "other_names": " ",
                            "gender_id": current_gender,
                            "dob": yob + "-" + mob + "-" + dob,
                            "head": head,
                            "marital": marital,
                            "current_village_id": current_village_id,
                            "card_issued": card_issued,
                            "audit_user_id": self._user._u.id
                        }
                        if nin is not False:
                            head_insuree_data["passport"] = nin
                        if r.get("Revenus") is not None and r.get("Revenus") != "":
                            head_insuree_data["income_level_id"] = int(r.get("Revenus"))+1 #+1 parce que les revenus
                            # en BD commencent a 1 et non 0
                        if r.get("Lien de parenté") is not None and r.get("Lien de parenté") != "":
                            head_insuree_data["relationship_id"] = int(r.get("Lien de parenté"))
                        if r.get("Catégories_professionnelles") is not None and r.get("Catégories_professionnelles") != "":
                            head_insuree_data["professional_situation"] = professional_situations.get(
                                int(r.get("Catégories_professionnelles"))
                                )
                        #    head_insuree_data["profession_id"] = int(r.get("Catégories_professionnelles"))
                        if r.get("Types de formation") is not None and r.get("Types de formation") != "":
                            head_insuree_data["education_id"] = int(r.get("Types de formation"))
                        jsonext = {}
                        jsonext.update({
                            "data": {
                                "head_insuree": head_insuree_data,
                                "family_level": "2" if marital == 2 else "1",
                                "location_id": current_village_id,
                                "family_type_id": "P" if marital == 2 else "H",
                            }
                        })
                        family_data = {
                            "head_insuree": head_insuree_data,
                            "family_level": "2" if marital == 2 else "1",
                            "location_id": current_village_id,
                            "family_type_id": "P" if marital == 2 else "H",
                            "json_ext": str(jsonext)
                        }
                        logger.info(f"family_data {family_data}" )
                        if not parent_family:
                            # Pas de famille donc on cree direct vu que les membre menages sont en ordre
                            # c'est le premier niveau de famille ici
                            parent_family = FamilyService(self._user).create_or_update(family_data)
                            logger.info(f"family created {parent_family}")
                            familycreated += 1
                            amount_family = False
                            if r.get("Cotisations_totales_ménages") is not None and r.get("Cotisations_totales_ménages") != "":
                                try:
                                    amount_family = Decimal(r.get("Cotisations_totales_ménages"))
                                except:
                                    logger.info("Could not parse the familly contribtion amount %s",
                                        r.get("Cotisations_totales_ménages"))
                            contribution_plan_code = False
                            current_contribution = False
                            policy_data = {}
                            if amount_family == 5000:
                                contribution_plan_code = "AMOS"
                            if amount_family == 3500:
                                contribution_plan_code = "AMOS1"
                            if amount_family == 2500:
                                contribution_plan_code = "AMOS2"
                            if amount_family == 2000:
                                contribution_plan_code = "AMOS3"
                            if amount_family == 1500:
                                contribution_plan_code = "AMOS4"
                            if amount_family == 0:
                                contribution_plan_code = "AMS"
                            if contribution_plan_code:
                                current_contribution = ContributionPlan.objects.filter(
                                    code="test"
                                ).first()
                                if current_contribution and contribution_plan_code:
                                    policy_data = {
                                        "enroll_date": today.date(),
                                        "start_date": today.date(),
                                        "status": Policy.STATUS_ACTIVE,
                                        "contribution_plan_id": str(current_contribution.id),
                                        "value": amount_family,
                                        "audit_user_id": self._user._u.id,
                                        "product_id": current_contribution.benefit_plan_id
                                    }
                            if marital != 2:
                                # On cree la police pour la famille si c'est pas poligame
                                # si c'est poligame c'est la sous famille qui aura la police
                                if current_contribution and contribution_plan_code:
                                    policy_data["family_id"] = parent_family.id
                                    logger.info("Creation police pour la famille %s", parent_family.id)
                                    policy_created = Policy.objects.create(**policy_data)
                                    logger.info("policy_created %s", policy_created.id)
                                    premium_data = {
                                        "audit_user_id": self._user._u.id,
                                        "receipt": f"receipt_{uuid4()}",
                                        "pay_date": today.date(),
                                        "pay_type": "B",
                                        "is_photo_fee": False,
                                        "amount": amount_family,
                                        "policy_id": policy_created.id
                                    }
                                    premium = Premium(**premium_data)
                                    created_premium = update_or_create_premium(premium, self._user)
                                    logger.info("premium created %s", created_premium)
                            if marital == 2:
                                #polygamous with should create 1 subfamily
                                jsonextsub = {}
                                jsonextsub.update({
                                    "family_data": {
                                        "family_level": "1",
                                        "location_id": current_village_id,
                                        "family_type_id": "H",
                                        "head_insuree_id": parent_family.head_insuree.id
                                    }
                                })
                                sub_family_data = {
                                    "family_level": "1",
                                    "location_id": current_village_id,
                                    "family_type_id": "H",
                                    "parent_id": parent_family.id,
                                    "json_ext": str(jsonextsub),
                                    "head_insuree_id": parent_family.head_insuree.id
                                }
                                logger.info("Creating subfamilly for family %s", parent_family.id)
                                sub_family = FamilyService(self._user).create_or_update(sub_family_data)
                                familycreated += 1
                                logger.info("sub family created %s", sub_family.id)
                                policy_data["family_id"] = sub_family.id
                                logger.info("Creation police pour la sous famille %s", sub_family.id)
                                policy_created = Policy.objects.create(**policy_data)
                                logger.info("sub family policy_created %s", policy_created)
                                premium_data = {
                                    "audit_user_id": self._user._u.id,
                                    "receipt": f"receipt_{uuid4()}",
                                    "pay_date": today.date(),
                                    "pay_type": "B",
                                    "is_photo_fee": False,
                                    "amount": amount_family,
                                    "policy_id": policy_created.id
                                }
                                premium = Premium(**premium_data)
                                created_premium = update_or_create_premium(premium, self._user)
                                logger.info("subfamily premium created %s", created_premium)
                        else:
                            # On ajoute l'assurée comme membre de la famille existante
                            head_insuree_data["family_id"] = parent_family.id
                            logger.info("creation assure simple pour la famille %s", parent_family)
                            insuree = InsureeService(self._user).create_or_update(head_insuree_data)
                            logger.info("Assuree cree %s", insuree)
                    logger.info("creation groupe d'itentification ok.......")
                logger.info("Fin du traitement d'import.......")
        except Exception as e:
            return InsureeImportExportService._get_general_error('FAILED TO IMPORT FILE: ', e)
        totals = {
            'sent': len(data_set),
            'created': familycreated,
        }
        errors = []
        success = True
        return success, totals, errors

class BankImportService:
    
    def __init__(self, user):
        self._user = user
        self._resource = InsureeResource(user)
    
    # Fonction pour parser les fichier BDC
    def parse_excel_bdc(self, file):
        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        transactions = []
        total_kmf = Decimal("0.00")

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if not row or all(cell is None for cell in row):
                continue
            if row_idx == 0:
                continue 

            try:
                chf_id = str(int(row[0])) if isinstance(row[0], float) else str(row[0]).strip()
                date = row[2]
                description = str(row[5]).strip()
                amount_raw = row[4]

                if not amount_raw:
                    raise ValueError("Montant manquant ou vide")

                # Nettoyage du montant
                amount_str = str(amount_raw).replace(",", ".").replace(" ", "")
                amount = Decimal(amount_str)    

                if amount > 0:
                    transactions.append({
                        "insuree_chf_id": chf_id,
                        "date": date.isoformat() if isinstance(date, datetime) else str(date),
                        "description": description,
                        "amount": str(amount),
                        "code_tp": "BDC",
                        "code_ext": f"bdc_{uuid4()}",
                        "code_receipt": f"receipt_{uuid4()}",
                        "label": description,
                        "fees": "0.00",
                        "amount_received": str(amount),
                        "date_payment": date.isoformat() if isinstance(date, datetime) else str(date),
                        "payment_origin": "Banque",
                        "payer_ref": chf_id,
                    })
                    total_kmf += amount
            except Exception as e:
                print(f"Erreur à la ligne {row_idx}: {row} - {e}")
                continue

        return {
            "transactions": transactions,
            "total_kmf": str(total_kmf),
            "count": len(transactions),
        }

    # Fonction pour parser les fichier Exim
    def parse_excel_exim(self, file):
        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        transactions = []
        total_kmf = Decimal("0.00")
        start_parsing = False

        for row in sheet.iter_rows(values_only=True):
            if not start_parsing:
                if row and any("Txn. Date" in str(cell) for cell in row if cell):
                    headers = [str(h).strip() if h else None for h in row]
                    try:
                        date_idx = headers.index("Txn. Date")
                        desc_idx = headers.index("Description")
                        credit_idx = headers.index("Credit")
                        reference_idx = headers.index("Txn.Ref No")
                    except ValueError as e:
                        raise Exception("Colonnes attendues non trouvées dans le fichier (Txn. Date, Description, Credit)") from e
                    start_parsing = True
                    continue

            if start_parsing:
                if row and str(row[1]).startswith("Opening Balance"):
                    break

                try:
                    credit_val = row[credit_idx]
                    if credit_val and Decimal(credit_val) > 0:
                        ref = str(row[reference_idx]) if row[reference_idx] else f"ref_{uuid4()}"
                        date_str = row[date_idx]
                        if isinstance(date_str, datetime):
                            date_formatted = date_str.date().isoformat()
                        else:
                            date_formatted = datetime.strptime(date_str, "%b %d, %Y").date().isoformat()
                        transactions.append({
                            "date": date_formatted,
                            "description": str(row[desc_idx]),
                            "amount": str(Decimal(credit_val)),
                            "insuree_chf_id": ref,
                            "code_ext": ref,
                            "label": str(row[desc_idx]),
                            "code_tp": "EXIM",
                            "code_receipt": f"receipt_{uuid4()}",
                            "fees": "0.00",
                            "amount_received": str(Decimal(credit_val)),
                            "date_payment": date_formatted,
                            "payment_origin": "Banque",
                            "payer_ref": ref,
                        })
                        total_kmf += Decimal(credit_val)
                except Exception as e:
                    print(f"Erreur parsing ligne: {row} - {e}")
                    continue

        return {
            "transactions": transactions,
            "total_kmf": str(total_kmf),
            "count": len(transactions),
        }
    
    def parse_date(self, date_str):
        for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%b %d, %Y", "%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        raise ValueError(f"Format de date non reconnu: {date_str}")


    def log_invoice_event(self, user, invoice, event_type, message):
        event = InvoiceEvent(
            invoice=invoice,
            event_type=event_type,
            message=message,
        )
        event.save(username=user.username)

    def get_invoice_by_code(self, code):
        try:
            return Invoice.objects.get(code=code, is_deleted=False)
        except Invoice.DoesNotExist:
            return None
        
    def find_invoice(self, chf_id):
        try:
            insuree = Insuree.objects.get(chf_id=chf_id, validity_to__isnull=True)
            family = Family.objects.get(head_insuree=insuree, validity_to__isnull=True)
            return Invoice.objects.filter(subject_id=str(family.id), is_deleted=False).order_by("date_invoice").first()
        except (Insuree.DoesNotExist, Family.DoesNotExist):
            return None

    def reconcile_bank_transaction(self, tx):
        chf_id = tx["insuree_chf_id"]
        if not chf_id:
            raise Exception("Insuree CHFID manquant")

        invoice = self.find_invoice(chf_id)
        if not invoice:
            raise Exception(f"Aucune facture trouvée pour le chf_id {chf_id}")

        if invoice.status in [Invoice.Status.PAID, Invoice.Status.CANCELLED]:
            raise Exception(f"Facture déjà payée ou annulée: {invoice.code}")

        payment_date = tx.get("date")
        if not payment_date:
            raise Exception("Date de paiement manquante")
        
        try:
            if isinstance(payment_date, datetime):
                payment_date = payment_date.date()
            elif isinstance(payment_date, str):
                payment_date = self.parse_date(payment_date)
        except ValueError:
            raise Exception(f"Format de date non reconnu: {payment_date}")

        subject_type = ContentType.objects.get_for_model(invoice)

        amount_received = Decimal(tx["amount_received"])
        code_ext = tx.get("code_ext") or f"pay_{uuid4()}"
        code_tp = tx.get("code_tp") or "Banque"
        code_receipt = tx.get("code_receipt") or f"receipt_{uuid4()}"
        label = tx.get("label") or f"Paiement pour {invoice.code}"
        reconciliation_status = PaymentInvoice.ReconciliationStatus.RECONCILIATED
        fees = Decimal(tx.get("fees", "0.00"))
        payment_origin = tx.get("payment_origin") or "Banque"
        payer_ref = tx.get("payer_ref") or chf_id

        payment_invoice = PaymentInvoice(
            code_ext=code_ext,
            code_tp=code_tp,
            code_receipt=code_receipt,
            label=label,
            reconciliation_status=reconciliation_status,
            amount_received=amount_received,
            fees=fees,
            date_payment=payment_date,
            payment_origin=payment_origin,
            payer_ref=payer_ref,
            payer_name=payer_ref
        )
        payment_invoice.save(username=self._user.username)

        detail_payment = DetailPaymentInvoice(
            payment=payment_invoice,
            subject_type=subject_type,
            subject_id=str(invoice.uuid),
            status=DetailPaymentInvoice.DetailPaymentStatus.ACCEPTED,
            fees=fees,
            amount=amount_received,
            reconcilation_id=f"recon_{uuid4()}",
            reconcilation_date=payment_date,
        )
        detail_payment.save(username=self._user.username)

        if invoice.status != Invoice.Status.RECONCILIATED:
            invoice.status = Invoice.Status.RECONCILIATED
            invoice.save(username=self._user.username)
        self.log_invoice_event(
            user=self._user,
            invoice=invoice,
            event_type=InvoiceEvent.EventType.PAYMENT,
            message=f"Paiement reçu pour l'assuré {chf_id}, montant {amount_received} KMF pour la date du {payment_date.strftime('%d/%m/%Y')}"
        )
        
        premium = self.create_premium(chf_id, payment_invoice)
        return {
            "invoice_code": invoice.code,
            "payment_id": payment_invoice.id,
            "detail_payment_id": detail_payment.id,
            "premium_uuid": str(premium.uuid) if premium else None,
            "status": "RECONCILIATED",
            "amount": str(invoice.amount_total),
        }

    def create_premium(self, chf_id, data):
        try:
            insuree = Insuree.objects.get(chf_id=chf_id, validity_to__isnull=True)
            family = Family.objects.get(head_insuree=insuree, validity_to__isnull=True)
            policy = Policy.objects.filter(
                family=family, validity_to__isnull=True, status=Policy.STATUS_IDLE,
            ).order_by("start_date").first()

            if policy:
                payer = Payer.objects.filter(type='C').first()
                
                premium_data = {
                    "audit_user_id": self._user.id,
                    "receipt": data.code_receipt,
                    "pay_date": data.date_payment,
                    "pay_type": "B",
                    "is_photo_fee": False,
                    "amount": data.amount_received,
                    "policy": policy,
                    "payer": payer
                }
                
                premium = Premium(**premium_data)
                created = update_or_create_premium(premium, self._user)
                logger.info(f"Contribution créée avec succès pour police {policy.id}")
                return created
            else:
                logger.warning(f"Aucune police active trouvée pour la famille {family.id}")
        except Exception as e:
            logger.exception(f"Erreur lors de la création de contribution pour CHFID {chf_id}: {e}")
    